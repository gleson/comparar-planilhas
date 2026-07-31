"""Gravação de edições no arquivo original, por formato (sem pandas/numpy)."""

import csv
import shutil
from pathlib import Path

from .files import _read_xls, _read_xlsx, _read_ods, detect_csv, list_sheets


def _coerce(value: str):
    """Converte o texto editado para o tipo mais natural ao gravar em planilha."""
    if value == "":
        return None
    try:
        return int(value)
    except ValueError:
        pass
    try:
        return float(value)
    except ValueError:
        pass
    return value


def _backup(path: Path) -> None:
    bak = path.with_name(path.name + ".bak")
    if not bak.exists():
        shutil.copy2(path, bak)


def _apply_to_grid(rows: list[list[str]], offset: int, edits: list[dict]) -> None:
    for e in edits:
        r, c = e["row"] + offset, e["col"]
        while len(rows) <= r:
            rows.append([])
        while len(rows[r]) <= c:
            rows[r].append("")
        rows[r][c] = e["value"]


def apply_edits(path: Path, sheet: str | None, has_header: bool,
                edits: list[dict]) -> dict:
    """Aplica edições {row, col, value} no arquivo original.

    `row` é o índice 0-based na área de dados (sem contar o cabeçalho).
    Retorna {"saved_path": str, "warnings": [str]}.
    """
    ext = path.suffix.lower()
    offset = 1 if has_header else 0
    warnings: list[str] = []

    if ext == ".csv":
        _backup(path)
        encoding, delimiter = detect_csv(path)
        with open(path, newline="", encoding=encoding) as f:
            rows = [list(r) for r in csv.reader(f, delimiter=delimiter)]
        _apply_to_grid(rows, offset, edits)
        with open(path, "w", newline="", encoding=encoding) as f:
            csv.writer(f, delimiter=delimiter).writerows(rows)
        return {"saved_path": str(path), "warnings": warnings}

    if ext == ".xlsx":
        import openpyxl

        _backup(path)
        wb = openpyxl.load_workbook(path)
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
        for e in edits:
            ws.cell(row=e["row"] + offset + 1, column=e["col"] + 1,
                    value=_coerce(e["value"]))
        wb.save(path)
        return {"saved_path": str(path), "warnings": warnings}

    if ext == ".ods":
        # Regravado por completo (dados de todas as abas preservados; a
        # formatação visual é perdida).
        _backup(path)
        names = list_sheets(path)
        sheets = {name: _read_ods(path, name) for name in names}
        target = sheet if sheet in sheets else (names[0] if names else "Planilha1")
        grid = sheets.setdefault(target, [])
        _apply_to_grid(grid, offset, edits)
        _write_ods(path, sheets)
        warnings.append(
            "Arquivo .ods regravado: os dados de todas as abas foram "
            "preservados, mas a formatação visual foi perdida."
        )
        return {"saved_path": str(path), "warnings": warnings}

    if ext == ".xls":
        # Formato legado sem escrita confiável: salva uma cópia .xlsx ao lado
        # do original, que não é alterado.
        import openpyxl

        out_path = path.with_suffix(".xlsx")
        names = list_sheets(path)
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for name in names:
            ws = wb.create_sheet(name)
            grid = _read_xls(path, name)
            if name == (sheet if sheet in names else (names[0] if names else None)):
                _apply_to_grid(grid, offset, edits)
            for row in grid:
                ws.append([_coerce(v) for v in row])
        wb.save(out_path)
        warnings.append(
            "Arquivos .xls (formato legado) não podem ser regravados com "
            f"segurança: as edições foram salvas em '{out_path.name}', ao "
            "lado do original, que não foi alterado."
        )
        return {"saved_path": str(out_path), "warnings": warnings}

    raise ValueError(f"Formato não suportado para gravação: {ext}")


def _write_ods(path: Path, sheets: dict[str, list[list[str]]]) -> None:
    """Escreve um .ods a partir de um dicionário {aba: matriz de strings}."""
    from odf.opendocument import OpenDocumentSpreadsheet
    from odf.table import Table, TableRow, TableCell
    from odf.text import P

    doc = OpenDocumentSpreadsheet()
    for name, rows in (sheets or {"Planilha1": []}).items():
        table = Table(name=name)
        for row in rows:
            tr = TableRow()
            for val in row:
                num = _coerce(val)
                if isinstance(num, bool):
                    num = None
                if isinstance(num, (int, float)):
                    tc = TableCell(valuetype="float", value=repr(num))
                    tc.addElement(P(text=str(val)))
                else:
                    tc = TableCell()
                    tc.addElement(P(text=str(val) if val else ""))
                tr.addElement(tc)
            table.addElement(tr)
        doc.spreadsheet.addElement(table)
    doc.save(str(path))
