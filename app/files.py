"""Leitura de planilhas (csv, xls, xlsx, ods) sem depender de pandas/numpy.

Usa bibliotecas puras em Python (openpyxl, xlrd, odfpy) para que a instalação
funcione em qualquer versão de Python (Linux/Windows) sem compilador C.
"""

import csv
import subprocess
import sys
from datetime import date, datetime, time
from pathlib import Path

SUPPORTED_EXTS = {".csv", ".xls", ".xlsx", ".ods"}

CSV_ENCODINGS = ["utf-8-sig", "utf-8", "cp1252", "latin-1"]

# Limite de repetição ao expandir células/linhas vazias de arquivos .ods
# (o formato costuma marcar milhares de colunas/linhas vazias no fim).
_ODS_MAX_REPEAT = 4096


def file_type(path: str) -> str:
    return Path(path).suffix.lower().lstrip(".")


def validate_path(path: str) -> Path:
    p = Path(path).expanduser()
    if not p.is_file():
        raise ValueError(f"Arquivo não encontrado: {p}")
    if p.suffix.lower() not in SUPPORTED_EXTS:
        raise ValueError(
            f"Formato não suportado: '{p.suffix}'. Use csv, xls, xlsx ou ods."
        )
    return p


def pick_file_dialog() -> str | None:
    """Abre um diálogo nativo de seleção de arquivo (tkinter) em subprocesso.

    Retorna o caminho escolhido, "" se cancelado, ou None se o diálogo
    não estiver disponível (sem tkinter/ambiente gráfico).
    """
    code = (
        "import tkinter as tk\n"
        "from tkinter import filedialog\n"
        "r = tk.Tk(); r.withdraw(); r.attributes('-topmost', True)\n"
        "print(filedialog.askopenfilename(title='Selecione a planilha',"
        " filetypes=[('Planilhas', '*.csv *.xls *.xlsx *.ods'), ('Todos', '*.*')]))"
    )
    try:
        proc = subprocess.run(
            [sys.executable, "-c", code],
            capture_output=True, text=True, timeout=600,
        )
    except Exception:
        return None
    if proc.returncode != 0:
        return None
    return proc.stdout.strip()


def detect_csv(path: Path) -> tuple[str, str]:
    """Detecta (encoding, delimitador) de um csv."""
    raw = path.read_bytes()
    encoding = "utf-8"
    for enc in CSV_ENCODINGS:
        try:
            raw.decode(enc)
            encoding = enc
            break
        except UnicodeDecodeError:
            continue
    sample = raw.decode(encoding, errors="replace")[:64 * 1024]
    try:
        delimiter = csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
    except csv.Error:
        delimiter = ";" if sample.count(";") > sample.count(",") else ","
    return encoding, delimiter


def fmt_cell(v) -> str:
    """Converte um valor de célula para o texto exibido."""
    if v is None:
        return ""
    if isinstance(v, bool):
        return "VERDADEIRO" if v else "FALSO"
    if isinstance(v, int):
        return str(v)
    if isinstance(v, float):
        if v != v:  # NaN
            return ""
        if v.is_integer() and abs(v) < 1e15:
            return str(int(v))
        return repr(v)
    if isinstance(v, datetime):
        if v.hour == v.minute == v.second == 0 and v.microsecond == 0:
            return v.strftime("%d/%m/%Y")
        return v.strftime("%d/%m/%Y %H:%M:%S")
    if isinstance(v, date):
        return v.strftime("%d/%m/%Y")
    if isinstance(v, time):
        return v.strftime("%H:%M:%S")
    return str(v)


# ---------------------------------------------------------------------------
# Listagem de abas
# ---------------------------------------------------------------------------

def list_sheets(path: Path) -> list[str]:
    ext = path.suffix.lower()
    if ext == ".csv":
        return []
    if ext == ".xlsx":
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True)
        try:
            return list(wb.sheetnames)
        finally:
            wb.close()
    if ext == ".xls":
        import xlrd
        book = xlrd.open_workbook(path, on_demand=True)
        return [str(n) for n in book.sheet_names()]
    if ext == ".ods":
        from odf.opendocument import load
        from odf.table import Table
        doc = load(str(path))
        return [t.getAttribute("name") for t in doc.getElementsByType(Table)]
    return []


# ---------------------------------------------------------------------------
# Leitura por formato
# ---------------------------------------------------------------------------

def _read_xlsx(path: Path, sheet: str | None) -> list[list[str]]:
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb[wb.sheetnames[0]]
        return [[fmt_cell(c) for c in row] for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def _xls_cell(book, cell) -> str:
    import xlrd
    t, v = cell.ctype, cell.value
    if t in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
        return ""
    if t == xlrd.XL_CELL_TEXT:
        return v
    if t == xlrd.XL_CELL_NUMBER:
        return fmt_cell(v)
    if t == xlrd.XL_CELL_DATE:
        try:
            y, mo, d, h, mi, s = xlrd.xldate_as_tuple(v, book.datemode)
            if (y, mo, d) == (0, 0, 0):
                return fmt_cell(time(h, mi, s))
            return fmt_cell(datetime(y, mo, d, h, mi, s))
        except Exception:
            return fmt_cell(v)
    if t == xlrd.XL_CELL_BOOLEAN:
        return "VERDADEIRO" if v else "FALSO"
    if t == xlrd.XL_CELL_ERROR:
        return xlrd.error_text_from_code.get(v, "#ERRO")
    return str(v)


def _read_xls(path: Path, sheet: str | None) -> list[list[str]]:
    import xlrd
    book = xlrd.open_workbook(path)
    sh = book.sheet_by_name(sheet) if sheet and sheet in book.sheet_names() \
        else book.sheet_by_index(0)
    rows = []
    for r in range(sh.nrows):
        rows.append([_xls_cell(book, sh.cell(r, c)) for c in range(sh.ncols)])
    return rows


def _ods_cell_text(tc) -> str:
    from odf import teletype
    return teletype.extractText(tc)


def _read_ods(path: Path, sheet: str | None) -> list[list[str]]:
    from odf.opendocument import load
    from odf.table import Table, TableRow, TableCell

    doc = load(str(path))
    tables = doc.getElementsByType(Table)
    if not tables:
        return []
    target = tables[0]
    if sheet is not None:
        for t in tables:
            if t.getAttribute("name") == sheet:
                target = t
                break

    def repeat(el, attr):
        try:
            return max(1, int(el.getAttribute(attr) or 1))
        except (TypeError, ValueError):
            return 1

    rows: list[list[str]] = []
    pending_rows = 0
    for tr in target.getElementsByType(TableRow):
        rep_r = repeat(tr, "numberrowsrepeated")
        cells: list[str] = []
        pending_cells = 0
        for tc in tr.getElementsByType(TableCell):
            rep_c = repeat(tc, "numbercolumnsrepeated")
            val = _ods_cell_text(tc)
            if val == "":
                pending_cells += rep_c            # adia células vazias (podem ser fim de linha)
            else:
                if pending_cells:
                    cells.extend([""] * min(pending_cells, _ODS_MAX_REPEAT))
                    pending_cells = 0
                cells.extend([val] * min(rep_c, _ODS_MAX_REPEAT))
        if not cells:
            pending_rows += rep_r                 # linha vazia: adia (pode ser fim do arquivo)
            continue
        if pending_rows:
            rows.extend([[] for _ in range(min(pending_rows, _ODS_MAX_REPEAT))])
            pending_rows = 0
        for _ in range(min(rep_r, _ODS_MAX_REPEAT)):
            rows.append(list(cells))
    return rows


def load_table(path: Path, sheet: str | None) -> list[list[str]]:
    """Carrega uma aba como matriz de strings, com linhas de mesmo comprimento."""
    ext = path.suffix.lower()
    if ext == ".csv":
        encoding, delimiter = detect_csv(path)
        with open(path, newline="", encoding=encoding) as f:
            rows = [list(r) for r in csv.reader(f, delimiter=delimiter)]
    elif ext == ".xlsx":
        rows = _read_xlsx(path, sheet)
    elif ext == ".xls":
        rows = _read_xls(path, sheet)
    elif ext == ".ods":
        rows = _read_ods(path, sheet)
    else:
        raise ValueError(f"Formato não suportado: {ext}")

    n_cols = max((len(r) for r in rows), default=0)
    for r in rows:
        r.extend([""] * (n_cols - len(r)))
    # Remove linhas totalmente vazias no fim
    while rows and all(c == "" for c in rows[-1]):
        rows.pop()
    return rows
