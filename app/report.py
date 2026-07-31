"""Exportação do relatório de diferenças em xlsx com destaques."""

import io
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill

FILL_DIFF = PatternFill("solid", fgColor="FFEB9C")    # âmbar: célula diferente
FILL_ONLY_A = PatternFill("solid", fgColor="FFC7CE")  # vermelho: linha só em A
FILL_ONLY_B = PatternFill("solid", fgColor="C6EFCE")  # verde: linha só em B
FILL_HEADER = PatternFill("solid", fgColor="D9E1F2")
BOLD = Font(bold=True)


def build_report(diff: dict, meta_a: dict, meta_b: dict) -> bytes:
    """Gera o xlsx do relatório: aba Resumo + abas A e B com destaques."""
    wb = openpyxl.Workbook()
    headers = diff["headers"]
    has_header = diff["has_header"]
    s = diff["summary"]

    ws = wb.active
    ws.title = "Resumo"
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 60
    rows = [
        ("Relatório de comparação de planilhas", ""),
        ("Gerado em", datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
        ("Planilha A", f"{meta_a['path']}" + (f" — aba: {meta_a['sheet']}" if meta_a.get("sheet") else "")),
        ("Planilha B", f"{meta_b['path']}" + (f" — aba: {meta_b['sheet']}" if meta_b.get("sheet") else "")),
        ("", ""),
        ("Células com valores diferentes", s["diff_cells"]),
        ("Linhas presentes só em A", s["rows_only_a"]),
        ("Linhas presentes só em B", s["rows_only_b"]),
        ("Linhas de dados em A", s["rows_a"]),
        ("Linhas de dados em B", s["rows_b"]),
    ]
    for r in rows:
        ws.append(r)
    ws["A1"].font = BOLD

    ws.append(("", ""))
    ws.append(("Diferenças (célula a célula)", ""))
    ws[f"A{ws.max_row}"].font = BOLD
    ws.append(("Linha (arquivo)", "Coluna", "Valor em A", "Valor em B"))
    for cell in ws[ws.max_row]:
        cell.font = BOLD
        cell.fill = FILL_HEADER
    offset = 2 if has_header else 1
    for row in diff["rows"]:
        if row["status"] == "only_a":
            ws.append((f"A: linha {row['a_idx'] + offset}", "—",
                       "linha inteira só em A", ""))
        elif row["status"] == "only_b":
            ws.append((f"B: linha {row['b_idx'] + offset}", "—", "",
                       "linha inteira só em B"))
        else:
            for c in row["diff_cols"]:
                ws.append((
                    f"A: {row['a_idx'] + offset} / B: {row['b_idx'] + offset}",
                    headers[c], row["a"][c], row["b"][c],
                ))

    for side, name in (("a", "Planilha A"), ("b", "Planilha B")):
        wsd = wb.create_sheet(name)
        if has_header:
            wsd.append(headers)
            for cell in wsd[1]:
                cell.font = BOLD
                cell.fill = FILL_HEADER
        for row in diff["rows"]:
            values = row[side]
            wsd.append(values if values is not None else [""] * diff["n_cols"])
            excel_row = wsd.max_row
            if row["status"] == "matched":
                for c in row["diff_cols"]:
                    wsd.cell(row=excel_row, column=c + 1).fill = FILL_DIFF
            elif row["status"] == f"only_{side}":
                fill = FILL_ONLY_A if side == "a" else FILL_ONLY_B
                for c in range(diff["n_cols"]):
                    wsd.cell(row=excel_row, column=c + 1).fill = fill

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
